"""Extract NRB BFS table C4 into the staging table (staging-and-review, P2 pattern).

Reads each acquired Excel file (from the raw lake via the manifest, or from a
local folder with --from-dir), parses table C4 with the pure layout module,
and UPSERTS the values into `nrb_bfs_staging` — the human-review holding area.
Nothing touches `observations` here; that is the promote step's job
(scripts/nrb_bfs.py) after review.

Upsert semantics per staging cell (bs month × indicator × class):
  - new cell                       -> inserted as review_status='pending'
  - same value again               -> untouched (idempotent re-runs)
  - CHANGED value                  -> value updated and status reset to
                                      'pending' — a source revision that must
                                      pass human review again before promotion.

Unrecognized row labels are reported and the run exits non-zero — we report,
we never guess (Prime Directive 7).

Run with `make nrb-bfs-extract`. Options:
    --dry-run        parse and report; write nothing (no database needed)
    --from-dir DIR   read .xlsx files from a local folder instead of the lake
    --limit N        process at most N files
"""

from __future__ import annotations

import argparse
import io
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import psycopg
from dotenv import load_dotenv

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from ingestion.common.io_utf8 import configure_stdout_utf8  # noqa: E402
from ingestion.common.raw_lake import RawLake  # noqa: E402
from ingestion.nrb.bfs_acquire import load_manifest  # noqa: E402
from ingestion.nrb.bfs_layout import BfsParseError, ParsedC4, parse_c4  # noqa: E402

C4_SCAN_ROWS = 80   # C4 ends around row 52; 80 leaves margin without a full scan
C4_SCAN_COLS = 7    # columns A..G


@dataclass
class SourceFile:
    name: str
    raw_ref: str      # raw-lake payload path, or 'local:<path>' in --from-dir mode
    source_url: str
    payload: bytes


def read_c4_matrix(payload: bytes) -> list[tuple[Any, ...]]:
    wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        if "C4" not in wb.sheetnames:
            raise BfsParseError("no sheet named 'C4' in workbook")
        ws = wb["C4"]
        return list(
            ws.iter_rows(min_row=1, max_row=C4_SCAN_ROWS, max_col=C4_SCAN_COLS, values_only=True)
        )
    finally:
        wb.close()


def iter_source_files(from_dir: str | None, limit: int | None) -> list[SourceFile]:
    files: list[SourceFile] = []
    if from_dir is not None:
        for p in sorted(Path(from_dir).glob("*.xls*")):
            files.append(SourceFile(p.name, f"local:{p}", f"file://{p}", p.read_bytes()))
    else:
        manifest = load_manifest()
        lake = RawLake.from_env()
        for name, entry in sorted(manifest.items()):
            payload = lake.backend.get(str(entry["payload_path"]))
            files.append(SourceFile(name, str(entry["payload_path"]),
                                    str(entry["source_url"]), payload))
    return files[:limit] if limit is not None else files


UPSERT_SQL = """
INSERT INTO nrb_bfs_staging
    (raw_ref, source_url, bs_year, bs_month, period_label, sheet,
     row_label, section, indicator_code, bfi_class, value, unit_code)
VALUES (%(raw_ref)s, %(source_url)s, %(bs_year)s, %(bs_month)s, %(period_label)s, 'C4',
        %(row_label)s, %(section)s, %(indicator_code)s, %(bfi_class)s, %(value)s, %(unit_code)s)
ON CONFLICT (bs_year, bs_month, indicator_code, bfi_class) DO UPDATE SET
    raw_ref      = EXCLUDED.raw_ref,
    source_url   = EXCLUDED.source_url,
    period_label = EXCLUDED.period_label,
    row_label    = EXCLUDED.row_label,
    section      = EXCLUDED.section,
    value        = EXCLUDED.value,
    unit_code    = EXCLUDED.unit_code,
    extracted_at = now(),
    -- a changed value on an already-reviewed cell is a REVISION: back to review
    review_status = CASE
        WHEN nrb_bfs_staging.value IS DISTINCT FROM EXCLUDED.value THEN 'pending'
        ELSE nrb_bfs_staging.review_status
    END
WHERE nrb_bfs_staging.value IS DISTINCT FROM EXCLUDED.value
   OR nrb_bfs_staging.raw_ref IS DISTINCT FROM EXCLUDED.raw_ref
"""


def stage(conn: psycopg.Connection[Any], src: SourceFile, parsed: ParsedC4) -> int:
    """Upsert one file's parsed values; returns rows actually written."""
    written = 0
    with conn.cursor() as cur:
        for v in parsed.values:
            cur.execute(
                UPSERT_SQL,
                {
                    "raw_ref": src.raw_ref,
                    "source_url": src.source_url,
                    "bs_year": parsed.bs_year,
                    "bs_month": parsed.bs_month,
                    "period_label": parsed.period_label,
                    "row_label": v.row_label,
                    "section": v.section,
                    "indicator_code": v.indicator_code,
                    "bfi_class": v.bfi_class,
                    "value": v.value,
                    "unit_code": v.unit_code,
                },
            )
            written += cur.rowcount
    return written


def main() -> None:
    configure_stdout_utf8()
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--from-dir", default=None)
    parser.add_argument("--limit", type=int, default=None)
    args = parser.parse_args()

    files = iter_source_files(args.from_dir, args.limit)
    if not files:
        raise SystemExit("no source files found — run `make nrb-bfs-acquire` first")

    conn: psycopg.Connection[Any] | None = None
    if not args.dry_run:
        load_dotenv()
        dsn = os.environ.get("DATABASE_URL", "").strip()
        if not dsn:
            raise SystemExit("DATABASE_URL is empty — fill it in .env (or use --dry-run)")
        conn = psycopg.connect(dsn)

    parsed_ok, values_total, written_total = 0, 0, 0
    all_unmatched: dict[str, str] = {}   # label -> first file seen in
    failures: list[str] = []
    months: list[tuple[int, int]] = []

    try:
        for src in files:
            try:
                parsed = parse_c4(read_c4_matrix(src.payload))
            except BfsParseError as exc:
                failures.append(f"{src.name}: {exc}")
                continue
            parsed_ok += 1
            values_total += len(parsed.values)
            months.append((parsed.bs_year, parsed.bs_month))
            for label in parsed.unmatched_labels:
                all_unmatched.setdefault(label, src.name)
            if conn is not None:
                written_total += stage(conn, src, parsed)
        if conn is not None:
            conn.commit()
    finally:
        if conn is not None:
            conn.close()

    months.sort()
    span = f"{months[0]} … {months[-1]}" if months else "—"
    mode = "DRY RUN — nothing written" if args.dry_run else f"{written_total} staging rows written"
    print(f"Parsed {parsed_ok}/{len(files)} file(s), {values_total} values, months {span}.")
    print(mode + ".")

    if failures:
        print("FILES THAT DID NOT PARSE (report, never guess):")
        for f in failures:
            print("  -", f)
    if all_unmatched:
        print("UNRECOGNIZED LABELS (add to bfs_layout.REGISTRY after review):")
        for label, fname in sorted(all_unmatched.items()):
            print(f"  - {label!r} (first seen in {fname})")
    if failures or all_unmatched:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
